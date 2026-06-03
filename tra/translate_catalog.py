import openpyxl
from deep_translator import GoogleTranslator
import time
import re

INPUT_FILE = r'c:\Users\fshos\OneDrive\my projects\tra\产品目录_ProductCatalog_Mar.6.xlsx'
OUTPUT_FILE = r'c:\Users\fshos\OneDrive\my projects\tra\ProductCatalog_Translated_EN.xlsx'

translator = GoogleTranslator(source='zh-CN', target='en')

def contains_chinese(text):
    if not isinstance(text, str):
        return False
    return bool(re.search(r'[\u4e00-\u9fff]', text))

def translate_text(text):
    if not isinstance(text, str) or not text.strip():
        return text
    if not contains_chinese(text):
        return text
    
    # Google Translate has a limit of ~5000 chars per request
    # Split long texts into chunks
    MAX_CHUNK = 4500
    if len(text) <= MAX_CHUNK:
        try:
            result = translator.translate(text)
            return result if result else text
        except Exception as e:
            print(f"  ERROR translating short text: {e}")
            return text
    
    # For long texts, split by double newline (paragraphs) and translate each
    parts = text.split('\n\n')
    translated_parts = []
    for part in parts:
        if contains_chinese(part):
            # further split if still too long
            if len(part) > MAX_CHUNK:
                lines = part.split('\n')
                translated_lines = []
                chunk = ''
                for line in lines:
                    if len(chunk) + len(line) + 1 < MAX_CHUNK:
                        chunk += line + '\n'
                    else:
                        if chunk:
                            try:
                                tr = translator.translate(chunk.strip())
                                translated_lines.append(tr if tr else chunk)
                            except Exception as e:
                                print(f"  ERROR in chunk: {e}")
                                translated_lines.append(chunk)
                            time.sleep(0.3)
                            chunk = line + '\n'
                        else:
                            # single line too long, truncate and translate
                            try:
                                tr = translator.translate(line[:MAX_CHUNK])
                                translated_lines.append(tr if tr else line)
                            except Exception as e:
                                translated_lines.append(line)
                if chunk:
                    try:
                        tr = translator.translate(chunk.strip())
                        translated_lines.append(tr if tr else chunk)
                    except Exception as e:
                        translated_lines.append(chunk)
                    time.sleep(0.3)
                translated_parts.append('\n'.join(translated_lines))
            else:
                try:
                    tr = translator.translate(part)
                    translated_parts.append(tr if tr else part)
                    time.sleep(0.2)
                except Exception as e:
                    print(f"  ERROR in paragraph: {e}")
                    translated_parts.append(part)
        else:
            translated_parts.append(part)
    return '\n\n'.join(translated_parts)

wb_in = openpyxl.load_workbook(INPUT_FILE)
wb_out = openpyxl.Workbook()
ws_in = wb_in.active
ws_out = wb_out.active
ws_out.title = ws_in.title if ws_in.title else 'ProductCatalog'

total_cells = ws_in.max_row * ws_in.max_column
processed = 0
translated_count = 0

print(f"Starting translation: {ws_in.max_row} rows x {ws_in.max_column} cols")
print("=" * 60)

for row_idx, row in enumerate(ws_in.iter_rows(), start=1):
    for cell in row:
        val = cell.value
        processed += 1
        
        if val is not None and isinstance(val, str) and contains_chinese(val):
            print(f"[{processed}/{total_cells}] Translating Row {cell.row}, Col {cell.column} (len={len(val)})...")
            translated_val = translate_text(val)
            ws_out.cell(row=cell.row, column=cell.column, value=translated_val)
            translated_count += 1
            time.sleep(0.25)  # be polite to the API
        else:
            ws_out.cell(row=cell.row, column=cell.column, value=val)
    
    if row_idx % 10 == 0:
        print(f"--- Completed row {row_idx}/{ws_in.max_row} ---")

# Copy column widths approximately
for col in ws_in.column_dimensions:
    ws_out.column_dimensions[col].width = ws_in.column_dimensions[col].width

# Copy row heights
for row in ws_in.row_dimensions:
    ws_out.row_dimensions[row].height = ws_in.row_dimensions[row].height

wb_out.save(OUTPUT_FILE)
print()
print("=" * 60)
print(f"DONE! Translated {translated_count} cells.")
print(f"Output saved to: {OUTPUT_FILE}")
